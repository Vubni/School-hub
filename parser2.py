import re
import asyncio
from openpyxl import load_workbook
from database.database import Database  # Импортируйте ваш класс Database

async def parse_excel_and_save_to_db(file_path: str):
    # Загрузка workbook
    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb['Временное с 8-13 сентября ']
    
    # Загрузка списка всех возможных предметов из листа "pred"
    pred_sheet = wb['pred']
    subjects = set()
    for row in range(1, pred_sheet.max_row + 1):
        subject = pred_sheet.cell(row=row, column=1).value
        if subject:
            subjects.add(str(subject).strip())
    
    # Словарь для преобразования дней недели в числовой формат
    days_map = {
        'ПОНЕДЕЛЬНИК': 1,
        'ВТОРНИК': 2,
        'СРЕДА': 3,
        'ЧЕТВЕРГ': 4,
        'ПЯТНИЦА': 5,
        'СУББОТА': 6
    }
    
    current_day = None
    lessons_data = []
    processed_lessons = set()
    
    # Собираем информацию о всех классах
    class_columns = {}
    for col in range(3, sheet.max_column + 1):
        class_name = sheet.cell(row=3, column=col).value
        if class_name and class_name not in ['№', 'Каб/Предмет']:
            # Пробуем оба варианта парсинга названия класса
            match = re.match(r'(\d+)([А-Я]+)', str(class_name))
            if not match:
                match = re.match(r'(\d+)([А-ЯA-Z]+)', str(class_name).replace(' ', ''))
            if match:
                class_number = int(match.group(1))
                class_letter = match.group(2)
                class_columns[col] = (class_number, class_letter)
    
    # Проходим по всем строкам
    for row in range(4, sheet.max_row + 1):  # Начинаем с 4 строки, где начинаются данные
        # Определяем текущий день
        day_cell = sheet.cell(row=row, column=1).value
        if day_cell:
            day_str = str(day_cell).strip().upper()
            if day_str in days_map:
                current_day = days_map[day_str]
                continue
        
        if not current_day:
            continue
        
        # Получаем время урока (для проверки наличия времени)
        time_cell = sheet.cell(row=row, column=2).value
        
        # Для каждого класса
        for col, (class_number, class_letter) in class_columns.items():
            # Пропускаем пустые ячейки
            if not sheet.cell(row=row, column=col).value:
                continue
                
            # Определяем номер урока
            lesson_number = None
            
            # СПОСОБ 1: Ищем номер урока в текущей ячейке
            try:
                lesson_number = int(sheet.cell(row=row, column=col).value)
            except (ValueError, TypeError):
                pass
                
            # СПОСОБ 2: Ищем номер урока в ячейке слева
            if lesson_number is None and col > 3:
                try:
                    lesson_number = int(sheet.cell(row=row, column=col-1).value)
                except (ValueError, TypeError):
                    pass
            
            # СПОСОБ 3: Ищем номер урока в ячейке выше
            if lesson_number is None and row > 4:
                try:
                    lesson_number = int(sheet.cell(row=row-1, column=col).value)
                except (ValueError, TypeError):
                    pass
            
            # Если номер урока не найден, пропускаем эту ячейку
            if lesson_number is None:
                continue
                
            # Ищем предмет в текущей и соседних ячейках
            title = None
            search_cells = [
                (row, col),      # Текущая ячейка
                (row, col+1),    # Справа
                (row+1, col),    # Снизу
                (row+1, col+1),  # Снизу справа
                (row-1, col),    # Сверху
                (row-1, col+1),  # Сверху справа
            ]
            
            for r, c in search_cells:
                if 1 <= r <= sheet.max_row and 1 <= c <= sheet.max_column:
                    cell_value = sheet.cell(row=r, column=c).value
                    if cell_value and any(subject in str(cell_value) for subject in subjects):
                        title = str(cell_value).strip()
                        break
            
            # Если нашли предмет, добавляем урок
            if title:
                lesson_id = f"{class_number}_{class_letter}_{current_day}_{lesson_number}"
                
                if lesson_id not in processed_lessons:
                    lessons_data.append((
                        class_number, class_letter, current_day, title, lesson_number
                    ))
                    processed_lessons.add(lesson_id)
    
    # Сохранение в базу данных
    async with Database() as db:
        await db.execute("DELETE FROM lessons")
        insert_sql = """
            INSERT INTO lessons (class_number, class_letter, day_number, title, lesson_number)
            VALUES ($1, $2, $3, $4, $5)"""
        await db.executemany(insert_sql, lessons_data)

async def main():
    await parse_excel_and_save_to_db('Временное расписание уроков 2025-2026.xlsx')

if __name__ == '__main__':
    asyncio.run(main())