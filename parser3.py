import pandas as pd
from database.database import Database
import asyncio
from openpyxl import load_workbook
from openpyxl.styles import Alignment

async def export_users_to_excel(db: Database, filename: str = "users.xlsx"):
    """
    Экспортирует данные пользователей в Excel с разделением по классам и добавляет подписи
    """
    users = await db.execute_all("""
        SELECT name, surname, login, password, class_number, class_letter 
        FROM users 
        ORDER BY class_number, class_letter
    """)

    if not users:
        print("Не найдено пользователей для экспорта")
        return False

    df = pd.DataFrame(users)
    df = df.rename(columns={
        'name': 'Имя',
        'surname': 'Фамилия',
        'login': 'Логин',
        'password': 'Пароль',
        'class_number': 'Класс',
        'class_letter': 'Буква'
    })
    
    # Создаем Excel-файл
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for (class_num, class_letter), group in df.groupby(['Класс', 'Буква']):
            class_str = f"{class_num}{class_letter}"
            group.to_excel(writer, sheet_name=class_str, index=False)
            
        # Получаем workbook и обрабатываем каждый лист
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Определяем последнюю строку и объединяем столбцы
            last_row = sheet.max_row + 2
            merge_range = f"A{last_row}:F{last_row}"
            sheet.merge_cells(merge_range)
            sheet[merge_range.split(':')[0]] = 'Скачать приложение можно на сайте school-hub.ru'
            
            # Добавляем вторую подпись
            merge_range2 = f"A{last_row+1}:F{last_row+1}"
            sheet.merge_cells(merge_range2)
            sheet[merge_range2.split(':')[0]] = 'После авторизации сразу измените логин или пароль'
            
            merge_range2 = f"A{last_row+2}:F{last_row+2}"
            sheet.merge_cells(merge_range2)
            sheet[merge_range2.split(':')[0]] = 'чтобы никто больше не авторизовался под вашим именем!'
            
            # Выравниваем текст по центру
            for row in range(last_row, last_row+3):
                sheet[f"A{row}"].alignment = Alignment(horizontal='center')

    print(f"Данные успешно экспортированы в файл {filename}")
    return True

async def main():
    db = Database()
    async with db:
        await export_users_to_excel(db, "users1.xlsx")

asyncio.run(main())