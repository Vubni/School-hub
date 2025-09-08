from database.database import Database

# pip install openpyxl transliterate

import openpyxl
import random, asyncio
import string
from datetime import datetime
from transliterate import translit

class ExcelParser:
    def __init__(self, db):
        self.db = db
        self.date_formats = [
            '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y',
            '%Y.%m.%d', '%Y/%m/%d', '%Y-%m-%d',
            '%d.%m.%y', '%d/%m/%y', '%d-%m-%y'
        ]
    
    def generate_password(self, length=8):
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length))
    
    def transliterate_name(self, name):
        # Транслитерация кириллицы в латиницу
        try:
            return translit(name, 'ru', reversed=True)
        except:
            return name
    
    def generate_login(self, first_name, last_name):
        # Генерация логина в формате Имя_фамилия
        first_lat = self.transliterate_name(first_name).replace(' ', '_')
        last_lat = self.transliterate_name(last_name).replace(' ', '_')
        return f"{first_lat.capitalize()}_{last_lat.lower()}"
    
    def parse_date(self, date_str):
        # Парсинг даты из различных форматов
        if isinstance(date_str, datetime):
            return date_str.date()
        
        for fmt in self.date_formats:
            try:
                return datetime.strptime(str(date_str), fmt).date()
            except ValueError:
                continue
        return None
    
    def parse_sheet_name(self, sheet_name):
        # Разделение номера класса и буквы
        if not sheet_name:
            return None, None
        
        class_number = ''.join(filter(str.isdigit, sheet_name))
        class_letter = ''.join(filter(str.isalpha, sheet_name)).upper()
        
        return class_number, class_letter
    
    async def process_excel(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name in workbook.sheetnames:
            class_number, class_letter = self.parse_sheet_name(sheet_name)
            if not class_number:
                continue
            
            sheet = workbook[sheet_name]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):  # предполагаем заголовок в первой строке
                if not row[0]:  # пропускаем пустые строки
                    continue
                
                # Парсинг ФИО
                fio_parts = str(row[0]).split()
                if len(fio_parts) < 2:
                    continue
                
                surname = fio_parts[0]
                name = fio_parts[1]
                middle_name = fio_parts[2] if len(fio_parts) > 2 else None
                
                # Парсинг даты рождения
                birthday = self.parse_date(row[1]) if row[1] else None
                
                # Генерация логина и пароля
                login = self.generate_login(name, surname[:7])
                password = self.generate_password()
                
                # Вставка в базу данных
                await self.db.execute(
                    """INSERT INTO users (name, surname, middle_name, birthday, 
                    class_number, class_letter, login, password)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8)""",
                    (name, surname, middle_name, birthday,
                    int(class_number), class_letter, login, password)
                )

# Пример использования
async def main():
    async with Database() as db:
        parser = ExcelParser(db)
        await parser.process_excel('users.xlsx')

if __name__ == '__main__':
    asyncio.run(main())